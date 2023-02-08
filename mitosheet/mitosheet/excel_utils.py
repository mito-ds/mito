


import string
from typing import Optional, Tuple, Union

from openpyxl import load_workbook

from mitosheet.errors import make_invalid_range_error


def get_excel_range_from_column_index(col_index: int) -> str:
    """
    Number to Excel-style column name, e.g., 1 = A:A, 26 = Z:Z, 27 = AA:AA, 703 = AAA:AAA.
    """
    return get_column_from_column_index(col_index) + ":" + get_column_from_column_index(col_index)

def get_column_from_column_index(col_index: int) -> str:
    """
    Number to Excel-style column name, e.g., 1 = A, 26 = Z, 27 = AA, 703 = AAA.
    """
    # Add 1 because Mito 0 indexes columns
    col_index = col_index + 1
    name = ''
    while col_index > 0:
        col_index, r = divmod (col_index - 1, 26)
        name = chr(r + ord('A')) + name
    return name

def get_col_and_row_indexes_from_range(range: str) -> Tuple[Tuple[int, int], Tuple[int, int]]:
    try:
        start_cell_address, end_cell_address = range.split(':')
        (start_col_idx, start_row_idx) = get_row_and_col_index_from_cell_address(start_cell_address)
        (end_col_idx, end_row_idx) = get_row_and_col_index_from_cell_address(end_cell_address)
    except:
        raise make_invalid_range_error(range, False)

    return ((min(start_col_idx, end_col_idx), min(start_row_idx, end_row_idx)), (max(start_col_idx, end_col_idx), max(start_row_idx, end_row_idx)))

def get_row_and_col_index_from_cell_address(cell_address: str) -> Tuple[int, int]:
    
    letters = ''
    numbers = ''
    for char in cell_address:
        if char.isalpha():
            letters += char
        else:
            numbers += char

    return (get_index_from_column(letters), int(numbers) - 1)

def get_index_from_column(col: str) -> int:
    num = 0
    for c in col:
        if c in string.ascii_letters:
            num = num * 26 + (ord(c.upper()) - ord('A')) + 1
    return num - 1


def get_table_range_from_upper_left_corner_value(file_path: str, sheet_name: str, value: Union[str, int, float, bool]) -> Optional[str]:
    """
    Given a string, this function will look through the excel tab sheet_name at the given
    file_path, and find the first instance of the value (look through column A first, then B, etc).

    If this value does not exist in the tab, will return None. 

    If the value does exist, then this function will walk down the column until it hits an empty row. Then,
    it will walk down the final row until it hits an empty column. Then it will return the range that 
    defines that rectangular of defined data (with value in the upper left corner).
    """
    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name]

    # We get the last defined rows, so we don't waste time searching data we don't need
    dimension = sheet.calculate_dimension()
    (min_search_col, min_search_row), (max_search_col, max_search_row) = get_col_and_row_indexes_from_range(dimension)
    # Unfortunately, openpyxl indexes from 1, so we add one to treat everything in that range
    (min_search_col, min_search_row), (max_search_col, max_search_row) = (min_search_col + 1, min_search_row + 1), (max_search_col + 1, max_search_row + 1)

    # Loop over the columns one by one to find where this value is set
    min_found_col_index, min_found_row_index = None, None
    for col in sheet.iter_cols(min_row=min_search_row, max_row=max_search_row, min_col=min_search_col, max_col=max_search_col):
        for cell in col:
            if cell.value == value:
                min_found_col_index, min_found_row_index = cell.column, cell.row
                break
        
        # As soon as we find something, stop looking
        if min_found_col_index is not None or min_found_row_index is not None:
            break

    if min_found_col_index is None or min_found_row_index is None:
        return None

    column = sheet[get_column_from_column_index(min_found_col_index - 1)] # We need to subtract 1 as we 0 index
    max_found_row_index = None
    for cell in column:
        if cell.row < min_found_row_index:
            continue
        
        if cell.value is None:
            max_found_row_index = cell.row - 1 # minus b/c this is one past the end
            break

    # If we looped over the entire column without ending, then we set the max row index
    # as the length of the entire column
    if max_found_row_index is None:
        max_found_row_index = len(column)

    # Then we find find where the rows are defined to
    max_found_col_index = None
    for row in sheet.iter_rows(min_row=max_found_row_index - 1, max_row=max_found_row_index - 1, min_col=min_found_col_index):
        for cell in row:
            if cell.value is None:
                max_found_col_index = cell.column - 1 # minus b/c this is one past the end
                break
    
    # Similarly, if we don't find any empty value in the defined cells, we set the max_col index
    # as the limit of the sheet
    if max_found_col_index is None:
        max_found_col_index = max_search_col

    return f'{get_column_from_column_index(min_found_col_index - 1)}{min_found_row_index}:{get_column_from_column_index(max_found_col_index - 1)}{max_found_row_index}'


def get_df_name_as_valid_sheet_name(df_name: str) -> str:
    return df_name[:31] # not more than 32 chars