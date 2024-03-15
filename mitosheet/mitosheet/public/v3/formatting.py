from typing import Any, List, Optional, Dict

from openpyxl.styles import Font, PatternFill
from openpyxl.styles import NamedStyle
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.worksheet import Worksheet
from pandas import DataFrame, ExcelWriter

from mitosheet.excel_utils import get_column_from_column_index
from mitosheet.is_type_utils import is_float_dtype, is_int_dtype
from mitosheet.types import (
    FC_BOOLEAN_IS_FALSE, FC_BOOLEAN_IS_TRUE, FC_DATETIME_EXACTLY,
    FC_DATETIME_GREATER, FC_DATETIME_GREATER_THAN_OR_EQUAL, FC_DATETIME_LESS,
    FC_DATETIME_LESS_THAN_OR_EQUAL, FC_DATETIME_NOT_EXACTLY, FC_EMPTY,
    FC_NOT_EMPTY, FC_NUMBER_EXACTLY,
    FC_NUMBER_GREATER, FC_NUMBER_GREATER_THAN_OR_EQUAL, FC_NUMBER_HIGHEST,
    FC_NUMBER_LESS, FC_NUMBER_LESS_THAN_OR_EQUAL, FC_NUMBER_LOWEST,
    FC_NUMBER_NOT_EXACTLY, FC_STRING_CONTAINS, FC_STRING_DOES_NOT_CONTAIN,
    FC_STRING_ENDS_WITH, FC_STRING_EXACTLY, FC_STRING_NOT_EXACTLY,
    FC_STRING_STARTS_WITH, FC_STRING_CONTAINS_CASE_INSENSITIVE)


# Object to map the conditional formatting operators to the excel formulas
CONDITION_TO_COMPARISON_FORMULA: Dict[str, str] = {
    FC_NUMBER_GREATER: '>',
    FC_NUMBER_LESS: '<',
    FC_NUMBER_NOT_EXACTLY: '<>',
    FC_NUMBER_GREATER_THAN_OR_EQUAL: '>=',
    FC_NUMBER_LESS_THAN_OR_EQUAL: '<=',
    FC_DATETIME_EXACTLY: '=',
    FC_DATETIME_NOT_EXACTLY: '<>',
    FC_DATETIME_GREATER: '>',
    FC_DATETIME_GREATER_THAN_OR_EQUAL: '>=',
    FC_DATETIME_LESS: '<',
    FC_DATETIME_LESS_THAN_OR_EQUAL: '<=',
}

def get_conditional_format_rule(
    filter_condition: str,
    fill: Optional[PatternFill],
    font: Optional[Font],
    filter_value: str,
    cell_range: str,
    column: str
) -> Optional[FormulaRule]:
    # Update the formulas for the string operators
    comparison = CONDITION_TO_COMPARISON_FORMULA.get(filter_condition)
    # If comparing dates, we need to use the DATEVALUE function
    if comparison is not None and 'datetime' in filter_condition:
        formula = [f'{cell_range}{comparison}DATEVALUE("{filter_value}")']
    elif comparison is not None:
        formula = [f'{cell_range}{comparison}{filter_value}']
    elif filter_condition == FC_NUMBER_EXACTLY:
        # Sometimes, users will enter a string for a number filter -- because it is just
        # displayed as an = on the frontend. This is a reasonably thing for them
        # to do -- so we take special care in this case to wrap the filter value in
        # quotes
        try:
            float(filter_value)
            formula = [f'{cell_range}={filter_value}']
        except ValueError:
            formula = [f'{cell_range}="{filter_value}"']
    elif filter_condition == FC_STRING_CONTAINS:
        formula = [f'NOT(ISERROR(FIND("{filter_value}",{cell_range})))']
    elif filter_condition == FC_STRING_CONTAINS_CASE_INSENSITIVE:
        formula = [f'NOT(ISERROR(SEARCH("{filter_value}",{cell_range})))']
    elif filter_condition == FC_STRING_DOES_NOT_CONTAIN:
        formula = [f'ISERROR(SEARCH("{filter_value}",{cell_range}))']
    elif filter_condition == FC_STRING_STARTS_WITH:
        formula = [f'LEFT({cell_range},LEN("{filter_value}"))="{filter_value}"']
    elif filter_condition == FC_STRING_ENDS_WITH:
        formula = [f'RIGHT({cell_range},LEN("{filter_value}"))="{filter_value}"']
    elif filter_condition == FC_BOOLEAN_IS_TRUE:
        formula = [cell_range]
    elif filter_condition == FC_BOOLEAN_IS_FALSE:
        formula = [f'NOT({cell_range})']
    elif filter_condition == FC_STRING_EXACTLY:
        formula = [f'EXACT({cell_range},"{filter_value}")']
    elif filter_condition == FC_STRING_NOT_EXACTLY:
        formula = [f'NOT(EXACT({cell_range},"{filter_value}"))']
    elif filter_condition == FC_EMPTY:
        formula = [f'ISBLANK({cell_range})']
    elif filter_condition == FC_NOT_EMPTY:
        formula = [f'NOT(ISBLANK({cell_range}))']
    elif filter_condition == FC_NUMBER_HIGHEST:
        formula = [f'{column}2>=LARGE({column}:{column},{filter_value})']
    elif filter_condition == FC_NUMBER_LOWEST:
        formula = [f'{column}2<=SMALL({column}:{column},{filter_value})']
    else: return None
    return FormulaRule(fill=fill, font=font, formula=formula)

def add_conditional_formats(
    conditional_formats: Optional[List[Any]],
    sheet: Worksheet,
    df: DataFrame
) -> None:
    if conditional_formats is None:
        return

    for conditional_format in conditional_formats:
        for filter in conditional_format.get('filters', []):
            # Create the conditional formatting color objects
            cond_fill = None
            cond_font = None
            if conditional_format.get('background_color') is not None:
                cond_fill = PatternFill(start_color=conditional_format['background_color'][1:], end_color=conditional_format['background_color'][1:], fill_type='solid')
            if conditional_format.get('font_color') is not None:
                cond_font = Font(color=conditional_format['font_color'][1:])
            if cond_fill is None and cond_font is None:
                continue
            
            # Add the conditional formatting rule to the sheet
            for column_header in conditional_format['columns']:
                column_index = df.columns.tolist().index(column_header)
                column = get_column_from_column_index(column_index)
                cell_range = f'{column}2:{column}{sheet.max_row}'
                column_conditional_rule = get_conditional_format_rule(
                    filter_condition=filter['condition'],
                    fill=cond_fill,
                    font=cond_font,
                    filter_value=f"{filter['value']}",
                    cell_range=cell_range,
                    column=column
                )
                if column_conditional_rule is not None:
                    sheet.conditional_formatting.add(cell_range, column_conditional_rule)


def add_number_formatting(
    number_formats: Optional[Dict[str, str]],
    sheet: Worksheet,
    df: DataFrame
) -> None:
    default_number_format_column_indexes = set(range(len(df.columns)))

    # For all the columns that have user defined number formats 
    # apply those
    if number_formats is not None:
        for column_header, number_format in number_formats.items():
            column_index = df.columns.tolist().index(column_header)
            column = get_column_from_column_index(column_index)
            cell_range = f'{column}2:{column}{sheet.max_row}'
            for cell in sheet[cell_range]:
                cell[0].number_format = number_format

            default_number_format_column_indexes.remove(column_index)

    # For all the float columns that have default number formats
    # and contain numbers, apply the default number format
    # which has comma separators and 2 decimal places
    column_headers = df.columns.tolist()
    for column_index in default_number_format_column_indexes:
        column_header = column_headers[column_index]
        dtype = str(df[column_header].dtype)
        column = get_column_from_column_index(column_index)
        cell_range = f'{column}2:{column}{sheet.max_row}'

        if is_float_dtype(dtype):
            for cell in sheet[cell_range]:
                cell[0].number_format = '#,##0.00'
        if is_int_dtype(dtype):
            for cell in sheet[cell_range]:
                cell[0].number_format = '#,##0'

def add_header_formatting_to_excel_sheet(
    writer: ExcelWriter,
    sheet_name: str,
    header_background_color: Optional[str]=None,
    header_font_color: Optional[str]=None
) -> None:
    """
    Adds formatting to the header row of the sheet_name, based on the formatting the user
    currently has applied in the frontend. 
    """
    workbook = writer.book
    sheet = workbook.get_sheet_by_name(sheet_name)
    
    # Add formatting to the header row   
    header_name = f"{sheet_name}_Header"
    header_format = NamedStyle(name=header_name)
    if header_font_color:
        # Remove the # from the color
        header_format.font = Font(color=header_font_color[1:])
    if header_background_color:
        # Remove the # from the color
        header_format.fill = PatternFill(start_color=header_background_color[1:], end_color=header_background_color[1:], fill_type="solid")
    
    # Only add format if there is a header color or background color
    has_header_formatting = header_background_color or header_font_color 
    if has_header_formatting:
        # Add named styles for the header rows to improve performance
        workbook.add_named_style(header_format)

        # Write the formatting to the sheet
        for col in range(1, sheet.max_column + 1):
            sheet.cell(row=1, column=col).style = header_name

def add_row_formatting_to_excel_sheet(
    writer: ExcelWriter,
    sheet_name: str,
    even_background_color: Optional[str]=None,
    even_font_color: Optional[str]=None,
    odd_background_color: Optional[str]=None,
    odd_font_color: Optional[str]=None
) -> None:
    """
    Adds formatting to the rows of the sheet_name, based on the formatting the user
    currently has applied in the frontend. 
    """
    workbook = writer.book
    sheet = workbook.get_sheet_by_name(sheet_name)

    # Add formatting to the rows
    even_name = f"{sheet_name}_Even"
    odd_name = f"{sheet_name}_Odd"
    even_format = NamedStyle(name=even_name)
    odd_format = NamedStyle(name=odd_name)

    # Remove the # from the colors and define the formatting objects
    if even_background_color:
        even_format.fill = PatternFill(start_color=even_background_color[1:], end_color=even_background_color[1:], fill_type="solid")
    if even_font_color:
        even_format.font = Font(color=even_font_color[1:])
    if odd_background_color:
        odd_format.fill = PatternFill(start_color=odd_background_color[1:], end_color=odd_background_color[1:], fill_type="solid")
    if odd_font_color:
        odd_format.font = Font(color=odd_font_color[1:])
    
    # Only add format if there is a background color or font color
    has_row_formatting = even_background_color or even_font_color or odd_background_color or odd_font_color
    if not has_row_formatting:
        return

    workbook.add_named_style(even_format)
    workbook.add_named_style(odd_format)

    for row in range(2, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            if row % 2 == 0:
                sheet.cell(row=row, column=col).style = even_name
            else:
                sheet.cell(row=row, column=col).style = odd_name


def add_formatting_to_excel_sheet(
    writer: ExcelWriter,
    sheet_name: str,
    df: DataFrame,
    header_background_color: Optional[str]=None,
    header_font_color: Optional[str]=None,
    even_background_color: Optional[str]=None,
    even_font_color: Optional[str]=None,
    odd_background_color: Optional[str]=None,
    odd_font_color: Optional[str]=None,
    conditional_formats: Optional[list]=None,
    number_formats: Optional[Dict[str, Any]]=None
) -> None:
    """
    Adds formatting to the sheet_name, based on the formatting the user
    currently has applied in the frontend. 

    NOTE: this is a Mito Pro feature.
    """
    workbook = writer.book
    sheet = workbook.get_sheet_by_name(sheet_name)

    add_header_formatting_to_excel_sheet(
        writer=writer,
        sheet_name=sheet_name,
        header_background_color=header_background_color,
        header_font_color=header_font_color
    )

    add_row_formatting_to_excel_sheet(
        writer=writer,
        sheet_name=sheet_name,
        even_background_color=even_background_color,
        even_font_color=even_font_color,
        odd_background_color=odd_background_color,
        odd_font_color=odd_font_color
    )

    add_conditional_formats(conditional_formats, sheet, df)

    add_number_formatting(number_formats, sheet, df)