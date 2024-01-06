

import csv
import os
from typing import Optional, Tuple, Union

import openpyxl
from openpyxl import load_workbook

from mitosheet.excel_utils import (get_col_and_row_indexes_from_range,
                                   get_column_from_column_index)


def get_table_range(
        file_path: str, 
        sheet_name: Optional[str]=None,
        upper_left_value: Optional[Union[str, int, float, bool]]=None, 
        # NOTE: we can't change the order of these first three parameters, for backwards compatibility
        sheet_index: Optional[int]=None, 
        upper_left_value_starts_with: Optional[Union[str, int, float, bool]]=None,
        upper_left_value_contains: Optional[Union[str, int, float, bool]]=None,
        bottom_left_corner_consecutive_empty_cells: Optional[int]=None, # TODO: at some point, we should rename this to num_empty_cells_in_final_row
        bottom_left_consecutive_empty_cells_in_first_column: Optional[int]=None,
        bottom_left_value: Optional[Union[str, int, float, bool]]=None, 
        bottom_left_value_starts_with: Optional[Union[str, int, float, bool]]=None,
        bottom_left_value_contains: Optional[Union[str, int, float, bool]]=None,
        row_entirely_empty: Optional[bool]=None,
        cumulative_number_of_empty_rows: Optional[int]=None,
        consecutive_number_of_empty_rows: Optional[int]=None,
        num_columns: Optional[int]=None
) -> Optional[str]:
    """
    Given a string, this function will look through the excel tab sheet_name at the given
    file_path and find a range that meets the conditions expressed by it's parameters.
    """
    workbook = load_workbook(file_path)

    if sheet_name is None and sheet_index is None:
        raise ValueError('Either sheet_name or sheet_index must be defined')
    elif sheet_name is not None and sheet_index is not None:
        raise ValueError('Only one of sheet_name or sheet_index can be defined')
    elif sheet_name is not None:
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.worksheets[sheet_index] # type : ignore

    # We get the last defined rows, so we don't waste time searching data we don't need
    dimension = sheet.calculate_dimension()
    (min_search_col, min_search_row), (max_search_col, max_search_row) = get_col_and_row_indexes_from_range(dimension)
    # Unfortunately, openpyxl indexes from 1, so we add one to treat everything in that range
    (min_search_col, min_search_row), (max_search_col, max_search_row) = (min_search_col + 1, min_search_row + 1), (max_search_col + 1, max_search_row + 1)

    # Loop over the columns one by one to find where this value is set
    min_found_col_index, min_found_row_index = None, None
    for col in sheet.iter_cols(min_row=min_search_row, max_row=max_search_row, min_col=min_search_col, max_col=max_search_col):
        for cell in col:
            if upper_left_value is not None and cell.value == upper_left_value:
                min_found_col_index, min_found_row_index = cell.column, cell.row
                break
            if upper_left_value_starts_with is not None and str(cell.value).startswith(str(upper_left_value_starts_with)):
                min_found_col_index, min_found_row_index = cell.column, cell.row
                break
            if upper_left_value_contains is not None and str(upper_left_value_contains) in str(cell.value):
                min_found_col_index, min_found_row_index = cell.column, cell.row
                break
        
        # As soon as we find something, stop looking
        if min_found_col_index is not None or min_found_row_index is not None:
            break

    if min_found_col_index is None or min_found_row_index is None:
        return None

    # Then we find find where the columns are defined to
    if num_columns is None:
        max_found_col_index = None
        for row in sheet.iter_rows(min_row=min_found_row_index, max_row=min_found_row_index, min_col=min_found_col_index):
            for cell in row:
                if cell.value is None:
                    max_found_col_index = cell.column - 1 # minus b/c this is one past the end
                    break
    else:
        max_found_col_index = min_found_col_index + num_columns - 1

    # Similarly, if we don't find any empty value in the defined cells, we set the max_col index
    # as the limit of the sheet
    if max_found_col_index is None:
        max_found_col_index = max_search_col

    # Then we find the max row index
    column = sheet[get_column_from_column_index(min_found_col_index - 1)] # We need to subtract 1 as we 0 index
    max_found_row_index = None

    # Check for number of empty cells conditions for rows
    if bottom_left_corner_consecutive_empty_cells is not None or row_entirely_empty is not None:
        for row in sheet.iter_rows(min_row=min_found_row_index, max_row=sheet.max_row+1, min_col=min_found_col_index, max_col=max_found_col_index):
            empty_count = sum([1 if c.value is None else 0 for c in row])
            if (bottom_left_corner_consecutive_empty_cells is not None and empty_count >= bottom_left_corner_consecutive_empty_cells) or \
                (row_entirely_empty is not None and empty_count >= len(row)):
                max_found_row_index = row[0].row - 1 # minus b/c this is one past the end
                break
            
    # Check for number of empty cells conditions for columns
    if max_found_row_index is None and bottom_left_consecutive_empty_cells_in_first_column is not None:
        empty_count = 0
        for cell in column:
            if cell.row <= min_found_row_index:
                continue

            if cell.value is None:
                empty_count += 1
            else:
                empty_count = 0

            # Check if we're at the end of the column, in which case the last cell is the max
            if cell.row == sheet.max_row:
                max_found_row_index = cell.row - empty_count # minus b/c we don't want to take the empty cells
                break

            if empty_count == bottom_left_consecutive_empty_cells_in_first_column:
                max_found_row_index = cell.row - empty_count # minus b/c we don't want to take the empty cells
                break
            

    if max_found_row_index is None and cumulative_number_of_empty_rows is not None:
        num_empty = 0
        for row in sheet.iter_rows(min_row=min_found_row_index, max_row=sheet.max_row+1, min_col=min_found_col_index, max_col=max_found_col_index):
            is_empty_row = all([c.value is None for c in row])
            if is_empty_row:
                num_empty += 1
            
            if num_empty >= cumulative_number_of_empty_rows:
                max_found_row_index = row[0].row - 1 # minus b/c this is one past the end
                break
            if row[0].row == sheet.max_row:
                max_found_row_index = row[0].row # Stop at the end as well
                break

    if max_found_row_index is None and consecutive_number_of_empty_rows is not None:
        num_empty = 0
        for row in sheet.iter_rows(min_row=min_found_row_index, max_row=sheet.max_row+1, min_col=min_found_col_index, max_col=max_found_col_index):
            is_empty_row = all([c.value is None for c in row])
            if is_empty_row:
                num_empty += 1
            else:
                num_empty = 0

            if num_empty >= consecutive_number_of_empty_rows:
                max_found_row_index = row[0].row - consecutive_number_of_empty_rows # minus b/c this is past the end, empty rows
                break
            if row[0].row == sheet.max_row:
                max_found_row_index = row[0].row # stop at the end as well
                break

    # Then check for other ending conditions
    if max_found_row_index is None:
        for cell in column:
            if cell.row <= min_found_row_index:
                continue
            
            # Stop as soon as we match the final value
            if bottom_left_value is not None and bottom_left_value == cell.value:
                max_found_row_index = cell.row
                break
            if bottom_left_value_starts_with is not None and str(cell.value).startswith(str(bottom_left_value_starts_with)):
                max_found_row_index = cell.row
                break
            if bottom_left_value_contains is not None and str(bottom_left_value_contains) in str(cell.value):
                max_found_row_index = cell.row
                break
            # NOTE: IF you add more conditions here, then add them to the condition below as well checking they are None
            # so that we can continue to handle the default case of finding the first empty cells
            if (bottom_left_value is None) and (bottom_left_value_starts_with is None) and (bottom_left_value_contains is None) \
                  and cell.value is None: 
                # NOTE: Check this condition last, as it's the final end condition, and for backwards compatibility
                # this means that the user is looking for the first empty cell. NOTE
                max_found_row_index = cell.row - 1 # minus b/c this is one past the end
                break

    # If we looped over the entire column without ending, then we set the max row index
    # as the length of the entire column
    if max_found_row_index is None:
        max_found_row_index = len(column)

    return f'{get_column_from_column_index(min_found_col_index - 1)}{min_found_row_index}:{get_column_from_column_index(max_found_col_index - 1)}{max_found_row_index}'


# We keep the old function name for backwards compatibility
get_table_range_from_upper_left_corner_value = get_table_range


def get_read_excel_params_from_range(range: str) -> Tuple[int, int, str]:
    ((start_col_index, start_row_index), (end_col_index, end_row_index)) = get_col_and_row_indexes_from_range(range)
    nrows = end_row_index - start_row_index
    usecols = get_column_from_column_index(start_col_index) + ':' + get_column_from_column_index(end_col_index)
    return start_row_index, nrows, usecols


def convert_csv_file_to_xlsx_file(csv_path: str, sheet_name: Union[str, int]) -> str:
    """Converts a CSV file to an XLSX file"""
    
    xlsx_path = os.path.splitext(csv_path)[0] + '_tmp.xlsx'

    # Loop over each row of the CSV and write it to the XLSX
    with open(csv_path, 'r') as csv_file:
        csv_reader = csv.reader(csv_file)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name if isinstance(sheet_name, str) else f'Sheet{sheet_name}'
        for row in csv_reader:
            ws.append(row)

        wb.save(xlsx_path)

    return xlsx_path